import sys
import uuid
from datetime import datetime
from PyQt5 import QtWidgets, QtCore, QtGui
import pandas as pd
from openpyxl import load_workbook, Workbook

try:
    import qdarkstyle
    THEME = True
except ImportError:
    THEME = False

ITEMS_FILE = r"C:\Users\mohit\Documents\Coding\Project\BILLING_SOFTWARE\items.xlsx"
BILLS_FILE = r"C:\Users\mohit\Documents\Coding\Project\BILLING_SOFTWARE\bills.xlsx"
LOGO_IMAGE = r"C:\Users\mohit\Documents\Coding\Project\BILLING_SOFTWARE\logo.jpg"

class BillingApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Billing System")
        self.setWindowIcon(QtGui.QIcon(LOGO_IMAGE))
        self.resize(900, 650)
        try:
            self.items_df = pd.read_excel(ITEMS_FILE)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to load inventory: {e}")
            self.items_df = pd.DataFrame(columns=["Barcode", "Name", "Quantity", "Price"])
        self.cart = []
        self.init_ui()

    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)

        # Logo at the top
        logo_label = QtWidgets.QLabel()
        pixmap = QtGui.QPixmap(LOGO_IMAGE)
        pixmap = pixmap.scaled(160, 80, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        logo_label.setPixmap(pixmap)
        logo_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(logo_label)

        # Search/Add Section
        search_layout = QtWidgets.QHBoxLayout()
        self.barcode_input = QtWidgets.QLineEdit()
        self.barcode_input.setPlaceholderText("Enter Barcode")
        self.barcode_input.setToolTip("Enter item barcode")
        self.qty_input = QtWidgets.QSpinBox()
        self.qty_input.setMinimum(1)
        self.qty_input.setToolTip("Quantity to add")
        self.add_btn = QtWidgets.QPushButton("Add Item")
        self.add_btn.setToolTip("Add item to cart")
        self.add_btn.clicked.connect(self.add_item)
        self.add_btn.setShortcut("Ctrl+S")
        search_layout.addWidget(self.barcode_input)
        search_layout.addWidget(QtWidgets.QLabel("Qty:"))
        search_layout.addWidget(self.qty_input)
        search_layout.addWidget(self.add_btn)
        layout.addLayout(search_layout)

        # Inventory search/filter
        filter_layout = QtWidgets.QHBoxLayout()
        self.search_input = QtWidgets.QLineEdit()
        self.search_input.setPlaceholderText("Search inventory by barcode or name...")
        self.search_input.textChanged.connect(self.filter_inventory)
        filter_layout.addWidget(QtWidgets.QLabel("Inventory Search:"))
        filter_layout.addWidget(self.search_input)
        layout.addLayout(filter_layout)

        # Table for cart
        self.table = QtWidgets.QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Barcode", "Name", "Qty", "Price"])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        layout.addWidget(self.table)

        # GST Checkbox
        gst_layout = QtWidgets.QHBoxLayout()
        self.gst_checkbox = QtWidgets.QCheckBox("Include GST (18%)")
        gst_layout.addWidget(self.gst_checkbox)
        gst_layout.addStretch()
        layout.addLayout(gst_layout)

        # Total & Buttons
        btn_layout = QtWidgets.QHBoxLayout()
        self.total_label = QtWidgets.QLabel("Total: ₹0.00")
        btn_layout.addWidget(self.total_label)
        self.gen_bill_btn = QtWidgets.QPushButton("Generate Bill")
        self.gen_bill_btn.setToolTip("Generate and save bill")
        self.gen_bill_btn.clicked.connect(self.generate_bill)
        btn_layout.addWidget(self.gen_bill_btn)
        self.export_btn = QtWidgets.QPushButton("Export Bill as CSV")
        self.export_btn.setToolTip("Export current bill to CSV file")
        self.export_btn.clicked.connect(self.export_bill_csv)
        btn_layout.addWidget(self.export_btn)
        self.backup_btn = QtWidgets.QPushButton("Backup Bills File")
        self.backup_btn.setToolTip("Backup all bills Excel file")
        self.backup_btn.clicked.connect(self.backup_bills)
        btn_layout.addWidget(self.backup_btn)
        layout.addLayout(btn_layout)

    def filter_inventory(self):
        text = self.search_input.text().lower()
        try:
            self.items_df = pd.read_excel(ITEMS_FILE)
        except Exception:
            self.items_df = pd.DataFrame(columns=["Barcode", "Name", "Quantity", "Price"])
        if text:
            mask = self.items_df["Barcode"].astype(str).str.lower().str.contains(text) | self.items_df["Name"].astype(str).str.lower().str.contains(text)
            filtered = self.items_df[mask]
        else:
            filtered = self.items_df
        # Optionally, show filtered inventory in a dialog or update a table (not shown in main UI for simplicity)
        # For now, just update the barcode input if only one match
        if len(filtered) == 1:
            self.barcode_input.setText(str(filtered.iloc[0]["Barcode"]))

    def add_item(self):
        barcode = self.barcode_input.text().strip()
        qty = self.qty_input.value()
        try:
            self.items_df = pd.read_excel(ITEMS_FILE)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to load inventory: {e}")
            return
        row = self.items_df[self.items_df['Barcode'].astype(str) == barcode]
        if row.empty:
            QtWidgets.QMessageBox.warning(self, "Error", "Item not found!")
            return
        item = row.iloc[0]
        name, price, stock_qty = item['Name'], item['Price'], item['Quantity']
        if qty > stock_qty:
            QtWidgets.QMessageBox.warning(self, "Error", f"Not enough stock! Available: {stock_qty}")
            return
        self.cart.append({'Barcode': barcode, 'Name': name, 'Qty': qty, 'Price': price})
        self.refresh_table()

    def refresh_table(self):
        self.table.setRowCount(0)
        total = 0
        for item in self.cart:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QtWidgets.QTableWidgetItem(str(item["Barcode"])))
            self.table.setItem(row, 1, QtWidgets.QTableWidgetItem(str(item["Name"])))
            self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(item["Qty"])))
            price = item["Price"] * item["Qty"]
            total += price
            self.table.setItem(row, 3, QtWidgets.QTableWidgetItem(f"₹{price:.2f}"))
        if self.gst_checkbox.isChecked():
            gst = total * 0.18
            total += gst
            self.total_label.setText(f"Total (GST included): ₹{total:.2f}")
        else:
            self.total_label.setText(f"Total: ₹{total:.2f}")

    def generate_bill(self):
        if not self.cart:
            QtWidgets.QMessageBox.warning(self, "Error", "Cart is empty!")
            return
        items = []
        total = 0
        for item in self.cart:
            total_item = item['Price'] * item['Qty']
            items.append([item['Barcode'], item['Name'], item['Qty'], item['Price'], total_item])
            total += total_item
        gst = 0
        if self.gst_checkbox.isChecked():
            gst = round(total * 0.18, 2)
        grand_total = total + gst

        bill_df = pd.DataFrame(items, columns=["Barcode", "Name", "Qty", "Unit Price", "Total"])
        bill_df.loc[len(bill_df)] = ["", "", "", "Subtotal", total]
        bill_df.loc[len(bill_df)] = ["", "", "", "GST (18%)", gst]
        bill_df.loc[len(bill_df)] = ["", "", "", "Grand Total", grand_total]

        bill_id = datetime.now().strftime('%Y%m%d_%H%M%S') + "_" + str(uuid.uuid4())[:6]
        sheet_name = f"Bill_{bill_id}"

        try:
            with pd.ExcelWriter(BILLS_FILE, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                bill_df.to_excel(writer, sheet_name=sheet_name, index=False)
            wb = load_workbook(BILLS_FILE)
            ws = wb[sheet_name]
            ws["A6"] = "Bill ID:"
            ws["B6"] = bill_id
            ws["A7"] = "Date:"
            ws["B7"] = datetime.now().strftime('%Y-%m-%d')
            ws["A8"] = "Time:"
            ws["B8"] = datetime.now().strftime('%H:%M:%S')
            wb.save(BILLS_FILE)
            # Reduce stock in inventory
            for item in self.cart:
                idx = self.items_df[self.items_df['Barcode'].astype(str) == item['Barcode']].index
                if not idx.empty:
                    self.items_df.at[idx[0], 'Quantity'] -= item['Qty']
            self.items_df.to_excel(ITEMS_FILE, index=False)
            QtWidgets.QMessageBox.information(self, "Success", f"Bill generated (Sheet: {sheet_name})")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to generate bill: {e}")
        finally:
            self.cart = []
            self.refresh_table()
            self.barcode_input.clear()
            self.qty_input.setValue(1)

    def export_bill_csv(self):
        if not self.cart:
            QtWidgets.QMessageBox.warning(self, "Error", "Cart is empty!")
            return
        items = []
        total = 0
        for item in self.cart:
            total_item = item['Price'] * item['Qty']
            items.append([item['Barcode'], item['Name'], item['Qty'], item['Price'], total_item])
            total += total_item
        gst = 0
        if self.gst_checkbox.isChecked():
            gst = round(total * 0.18, 2)
        grand_total = total + gst
        bill_df = pd.DataFrame(items, columns=["Barcode", "Name", "Qty", "Unit Price", "Total"])
        bill_df.loc[len(bill_df)] = ["", "", "", "Subtotal", total]
        bill_df.loc[len(bill_df)] = ["", "", "", "GST (18%)", gst]
        bill_df.loc[len(bill_df)] = ["", "", "", "Grand Total", grand_total]
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export Bill as CSV", "bill.csv", "CSV Files (*.csv)")
        if save_path:
            try:
                bill_df.to_csv(save_path, index=False)
                QtWidgets.QMessageBox.information(self, "Success", f"Bill exported to {save_path}")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Error", f"Failed to export bill: {e}")

    def backup_bills(self):
        import shutil
        try:
            backup_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Backup Bills Excel File", "bills_backup.xlsx", "Excel Files (*.xlsx)")
            if backup_path:
                shutil.copy2(BILLS_FILE, backup_path)
                QtWidgets.QMessageBox.information(self, "Success", f"Backup created at {backup_path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to backup: {e}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    if THEME:
        app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    window = BillingApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()